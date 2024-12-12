from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
import os

# Load the workbook and select the active sheet
file_path = r"C:\Users\kyleh\Downloads\Untitled spreadsheet.xlsx"  # Replace with the path to your Excel file
wb = load_workbook(file_path)
ws = wb.active

# Apply formatting to all rows in columns A and B
def format_datetime(cell):
    # Define the style if it doesn't already exist
    if 'CustomDateTime' not in wb.named_styles:
        date_style = NamedStyle(name="CustomDateTime", number_format="mm/dd/yy hh:mm:ss")
        wb.add_named_style(date_style)

    cell.style = "CustomDateTime"

# Determine the maximum row with data
max_row = ws.max_row

# Apply formatting to all rows in columns A and B
for row in range(2, max_row + 1):
    format_datetime(ws[f"A{row}"])
    format_datetime(ws[f"B{row}"])

# Add a new column (C) to the right of column B
ws.insert_cols(3)  # Insert a new column at position 3 (Column C)
ws["C1"] = "Average"  # Set header for the new column

# Add formula to calculate the average of (A - B) for all rows
for row in range(2, max_row + 1):
    ws[f"C{row}"] = f"=AVERAGE(A{row}-B{row})"

# Save the workbook
directory, filename = os.path.split(file_path)
updated_file_path = os.path.join(directory, "updated_" + filename)
wb.save(updated_file_path)

print(f"Spreadsheet updated and saved as {updated_file_path}")
